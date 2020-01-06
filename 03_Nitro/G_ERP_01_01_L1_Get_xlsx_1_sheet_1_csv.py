#!/usr/bin/env python
# coding: utf-8

# # 01-Trim off col and row of None from Original BOM and Save it as csv 

# In[1]:


import pandas as pd
import numpy as np
from openpyxl import load_workbook


# ### Get the wb to DataFrame
# * At this moment, the dataFrame BOM contains the cell of the wb
# * We can't read the value of the cell directly, jsut like there is a layer between the value and DataFrame BOM
# * Fetch worksheet by the sheet name given, if the name of sheet is given, take the first sheet.

# In[2]:


def get_xlsx(xlsx_data, sheetname = None):
    '''
    This func() takes 2 parameter.
    
    1st param.)  
    The path of the xlsx file.
    
    2nd param.) 
    It is the name of worksheet in the xlsm, appointing to a specific sheet.
    If the name is not given, the func() will takes the frist sheet in the xlsm.
    '''
    
    # get the xlsx file, read the data only not the fumula in cells of the xlsx.
    wb = load_workbook(xlsx_data, data_only = True)
    
    # if sheet name is not given, take the frist sheet of the xlsx file.
    if sheetname == None:
        sheetname = wb.sheetnames[0]
    
    # import the one of the data of worksheet from the xlsx as a DataFrame
    bom = pd.DataFrame(data = wb[sheetname])
   
    # Create an empty dataFrame in structure as bom
    new_bom = pd.DataFrame(index = bom.index, columns = bom.columns)
    
    # Fetch the values from bom and insert it to new_bom
    # We fetch the value and insert to the new_BOM, so we can directly read the values by dataFrame methods
    for col in bom.columns:
        for row in bom.index:
            new_bom[col][row] = bom[col][row].value
            
    return new_bom


# ### Delete column all with None
# #### Delete the column that is 
# *  new_BOM[column].count() == 0 

# In[3]:


def trim_None_col(bom):

    for col in bom:
        
        if bom[col].count() == 0:
            bom = bom.drop([col], axis = 1)
    
    return bom


# ### Delete row all with None
# #### The steps
# *  Check all the cell in the frist column, if the cell with None, get the index to the list "z"
# *  Check the rows with index in list "z", if the cells in the row are all in value None, then get the index of the row to list "row_want_delete"
# *  Delete the row by index in list "row_want_delete".

# In[4]:


def trim_None_row(bom):
    new_bom = bom
    z = []
    row_want_delete = []
    
    # Check the first cell with no value in first column
    #Get z
    z = bom[bom[0].values == None].index
   
    # Check if the row with first cell None is each of its cell all None
    # Get the index of row want to delete
    for index in z:
        if (bom.loc[index]).count().sum() == 0:
            row_want_delete.append(index)

    # Delete the rows
    new_bom = new_bom.drop(row_want_delete, axis = 0)
            
    return new_bom


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




